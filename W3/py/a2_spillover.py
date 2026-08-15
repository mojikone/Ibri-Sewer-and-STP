# -*- coding: utf-8 -*-
"""A2 — capacity-constrained growth reallocation (spillover model).
Total growth conserved (= R0 project total, user intent 2026-08-14): each settlement absorbs
its R0 growth up to its land ceiling (plots x OR 6.0 [GAP-5]); excess spills to neighbouring
settlements (boundary distance < 3 km) in proportion to remaining capacity, cascading outward.
Outputs: A2_spillover.csv, A2_spillover_chart.png."""
import zipfile, json, os, csv
import xml.etree.ElementTree as ET
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from shapely.geometry import Polygon, MultiPolygon
from shapely.ops import transform as shp_t
from rasterio.warp import transform as rio_transform
import openpyxl

W3 = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCR = r"C:\Users\mojtaba\AppData\Local\Temp\claude\D--Mojtaba-Renardet-2621-Ibri-Sewer-STP-Hydraulic-Claude\413b098f-38f9-4a93-9820-2b7c34af7f5e\scratchpad"
KMZ = r"D:\Mojtaba\Renardet\2621 Ibri Sewer STP\Data\Received\2621\inception report - R0\Final_Boundary_IBRI.kmz"
XLSX = r"D:\Mojtaba\Renardet\2621 Ibri Sewer STP\Data\Received\2621\inception report - R0\Ibri Sewer Demand R0 2026 08 03.xlsx"
NEIGH_DIST = 3000.0

# ---- geometry ----
ns = {"k": "http://www.opengis.net/kml/2.2"}
root = ET.fromstring(zipfile.ZipFile(KMZ).read("doc.kml").decode("utf-8", errors="replace"))
zg = {}
for pm in root.iter("{http://www.opengis.net/kml/2.2}Placemark"):
    nm_el = pm.find("k:name", ns)
    if nm_el is None: continue
    nm = nm_el.text.strip().upper()
    polys = []
    for poly in pm.iter("{http://www.opengis.net/kml/2.2}Polygon"):
        co = poly.find(".//k:outerBoundaryIs/k:LinearRing/k:coordinates", ns)
        if co is None: continue
        pts = [(float(t.split(",")[0]), float(t.split(",")[1])) for t in co.text.split() if len(t.split(",")) >= 2]
        if len(pts) >= 4: polys.append(Polygon(pts))
    if polys:
        g = MultiPolygon(polys) if len(polys) > 1 else polys[0]
        zg[nm] = zg[nm].union(g) if nm in zg else g
def to_utm(geom):
    def f(x, y, z=None):
        xs = list(x) if hasattr(x, "__iter__") else [x]
        ys = list(y) if hasattr(y, "__iter__") else [y]
        x2, y2 = rio_transform("EPSG:4326", "EPSG:32640", xs, ys)
        return (x2, y2)
    return shp_t(f, geom)
zg = {n: to_utm(g) for n, g in zg.items() if n != "0"}

# ---- ceilings from A1 ----
zones = {z["name"]: z for z in json.load(open(SCR + r"\zone_capacity.json")) if z["name"] != "0"}
# boundary-mismatch zones: ceiling unreliable (polygon misses plots) -> treat as uncapped
BAD_BOUNDARY = {"TANAM", "SATWAH", "AL MAKHTIBYAH", "USAYBUQ", "ASH SHIAB"}

# ---- projections ----
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Project Pop Settlements"]
hdr = [c.value for c in ws[1]]
yc = {int(str(h).replace("Pop ", "")): i for i, h in enumerate(hdr) if h and str(h).startswith("Pop ")}
proj = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[1]: continue
    nm = str(row[1]).strip().upper()
    proj[nm] = {y: float(row[i] or 0) for y, i in yc.items()}

names = [n for n in zg if n in proj and n in zones]
print("settlements in model:", len(names))
ceil_ = {n: (zones[n]["ceiling"] if n not in BAD_BOUNDARY else float("inf")) for n in names}

# ---- adjacency (boundary distance < NEIGH_DIST) ----
adj = {n: [] for n in names}
for i, a in enumerate(names):
    for b in names[i + 1:]:
        if zg[a].distance(zg[b]) < NEIGH_DIST:
            adj[a].append(b); adj[b].append(a)
print("IBRI neighbours:", adj.get("IBRI"))

# ---- spillover simulation ----
years = sorted(yc)
pop = {n: proj[n][years[0]] for n in names}
hist = {n: {} for n in names}
unhoused = {}
for k in range(len(years) - 1):
    y0, y1 = years[k], years[k + 1]
    for n in names:
        hist[n][y0] = pop[n]
    growth = {n: max(proj[n][y1] - proj[n][y0], 0.0) for n in names}
    # add own growth, collect overflow
    overflow = 0.0
    over_src = {}
    for n in names:
        pop[n] += growth[n]
        if pop[n] > ceil_[n]:
            over_src[n] = pop[n] - ceil_[n]
            pop[n] = ceil_[n]
    # spill: neighbours first (proportional to remaining capacity), then anywhere
    for src, ex in over_src.items():
        frontier = adj[src]
        for ring in range(4):
            # receivers need a KNOWN remaining capacity (unknown ceiling != infinite sink)
            caps = {m: ceil_[m] - pop[m] for m in frontier
                    if ceil_[m] != float("inf") and ceil_[m] - pop[m] > 1}
            total = sum(caps.values())
            if total > 0:
                take = min(ex, total)
                for m, c in caps.items():
                    pop[m] += take * c / total
                ex -= take
            if ex <= 1: break
            nxt = set()
            for m in frontier: nxt.update(adj[m])
            nxt.discard(src)
            frontier = [m for m in nxt if m not in frontier]
            if not frontier:
                frontier = [m for m in names if m != src]
        # residual beyond ALL known capacity: area is saturated -> track as unhoused
        if ex > 1: unhoused[y1] = unhoused.get(y1, 0.0) + ex
for n in names:
    hist[n][years[-1]] = pop[n]
print("unhoused surplus (area beyond capacity):")
for y in sorted(unhoused):
    if y in (2055, 2070, 2080, 2090, 2100) or y == min(unhoused):
        cum = sum(v for yy, v in unhoused.items() if yy <= y)
        print(f"  by {y}: cumulative {cum:,.0f} persons cannot be housed inside the boundary")

# ---- outputs ----
rows = []
for n in sorted(names, key=lambda m: -hist[m][2055]):
    r0_55, r0_00 = proj[n][2055], proj[n][2100]
    sp_55, sp_00 = hist[n][2055], hist[n][2100]
    rows.append({"settlement": n,
                 "ceiling": (round(ceil_[n]) if ceil_[n] != float("inf") else "uncapped*"),
                 "R0_2055": round(r0_55), "spill_2055": round(sp_55), "delta_2055": round(sp_55 - r0_55),
                 "R0_2100": round(r0_00), "spill_2100": round(sp_00), "delta_2100": round(sp_00 - r0_00)})
with open(os.path.join(W3, "analysis", "A2_spillover.csv"), "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=list(rows[0].keys())); w.writeheader(); w.writerows(rows)
print(f"{'settlement':<18}{'ceil':>9}{'R0-55':>9}{'sp-55':>9}{'d55':>8}{'R0-2100':>9}{'sp-2100':>9}{'d2100':>9}")
for r in rows[:12]:
    print(f"{r['settlement']:<18}{str(r['ceiling']):>9}{r['R0_2055']:>9}{r['spill_2055']:>9}{r['delta_2055']:>8}{r['R0_2100']:>9}{r['spill_2100']:>9}{r['delta_2100']:>9}")
tot_r0 = sum(proj[n][2055] for n in names); tot_sp = sum(hist[n][2055] for n in names)
print("2055 totals R0 vs spill:", round(tot_r0), round(tot_sp))

# chart: top-8 by spill_2055, R0 vs spillover
top = rows[:8]
x = np.arange(len(top)); wd = 0.36
fig, ax = plt.subplots(figsize=(9.5, 4.2))
ax.bar(x - wd / 2, [r["R0_2055"] / 1000 for r in top], wd, color="#9CC7E4", label="R0 allocation 2055")
ax.bar(x + wd / 2, [r["spill_2055"] / 1000 for r in top], wd, color="#0072B2", label="capacity-constrained 2055")
for i, r in enumerate(top):
    if isinstance(r["ceiling"], int):
        ax.plot([x[i] - 0.45, x[i] + 0.45], [r["ceiling"] / 1000] * 2, color="#C8342A", lw=1.4)
ax.set_xticks(x); ax.set_xticklabels([r["settlement"] for r in top], fontsize=8, rotation=20, ha="right")
ax.set_ylabel("population 2055 (thousands)", fontsize=9)
ax.legend(fontsize=8.5, frameon=False)
for s in ("top", "right"): ax.spines[s].set_visible(False)
ax.grid(axis="y", color="#DDDDDD", lw=0.7, zorder=0)
ax.set_title("2055 population: R0 vs capacity-constrained spillover (red line = land ceiling)", fontsize=11)
fig.savefig(os.path.join(W3, "img", "A2_spillover_2055.png"), dpi=170, bbox_inches="tight")
print("A2 outputs written")
