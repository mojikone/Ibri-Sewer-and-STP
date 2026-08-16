"""
A8 — transparent load-calculation workbook (W3/analysis/A8_load_calc.xlsx).

One sheet per question the user keeps having to re-ask:
  Inputs      every criteria value, editable, with its GUD-201 page ref
  Population  the 25 towns' NCSI series at the model years, and where it comes from
  Load        the full demand -> wastewater -> peak chain, one row per model year
  By Town 2055  corrected allocation: domestic per town, ND+Gov spread by
                non-residential plot area (A7 method)
  Saturation  land ceiling at OR 4.9 / 5.5 / 6.0 vs the projection

All numbers flow through live Excel formulas from the Inputs sheet, so changing
OR or LPCD re-computes everything. Data columns carry their source in a note row.
"""
from pathlib import Path

import geopandas as gpd
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(r"D:\Mojtaba\Renardet\2621 Ibri Sewer STP")
REPO = ROOT / "Hydraulic/Claude"
OUT = REPO / "W3/analysis/A8_load_calc.xlsx"

# ---------------------------------------------------------------- source data
towns = gpd.read_file(ROOT / "Hydraulic/SHP/Towns/Towns.shp")
towns = towns.sort_values("NAME_EN").reset_index(drop=True)
nonres = pd.read_csv(REPO / "W3/analysis/_nonres_area_by_town.csv")
towns = towns.merge(nonres, on="NAME_EN", how="left")
YEARS_SHOW = [2025, 2030, 2040, 2050, 2055, 2060, 2100]
ALL_YEARS = list(range(2023, 2101))

# ---------------------------------------------------------------- styles
F_TITLE = Font(name="Arial", size=13, bold=True)
F_H = Font(name="Arial", size=10, bold=True)
F_B = Font(name="Arial", size=10)
F_NOTE = Font(name="Arial", size=9, italic=True, color="666666")
F_LINK = Font(name="Arial", size=10, color="008000")      # green = other sheet
F_IN = Font(name="Arial", size=10, color="0000FF")         # blue = editable
FILL_H = PatternFill("solid", fgColor="D9E2F3")
FILL_IN = PatternFill("solid", fgColor="FFFF00")
FILL_TOT = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM = "#,##0"


def header(ws, row, labels, col0=1):
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=col0 + i, value=lab)
        c.font, c.fill, c.border = F_H, FILL_H, BOX
        c.alignment = Alignment(wrap_text=True, vertical="center",
                                horizontal="center")


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


wb = Workbook()

# ================================================================ READ ME
ws = wb.active
ws.title = "READ ME"
ws["B2"] = "Ibri sewer — load calculation, step by step"
ws["B2"].font = F_TITLE
lines = [
    "Every value is calculated live from the 'Inputs' sheet. Change an input "
    "(e.g. occupancy 6.0 -> 4.9) and the whole workbook follows.",
    "",
    "Sheets:",
    "  Inputs        all criteria values with their PAM-GUD-201 page refs. Yellow cells are editable.",
    "  Population    the 25 project towns, NCSI series at the model years. Source and method explained on the sheet.",
    "  Load          water demand -> wastewater -> infiltration -> Qadf -> peak factor -> peak flow, one row per model year.",
    "  By Town 2055  the same load split per town. Non-domestic + governmental volume is placed on non-residential",
    "                plot area, not spread per person (GUD-201 §7.3.2-7.3.3 — see A7 note).",
    "  Saturation    how much population the plots can hold at a given occupancy, and when the projection hits it.",
    "",
    "Colour code:  yellow fill = input, edit these   |   green text = pulled from another sheet   |   black = formula or data.",
    "",
    "Three cautions:",
    "  1. Occupancy 6.0 and 1 dwelling/plot are unconfirmed [GAP-5]. Measured anchor: ~7.0 persons per built plot.",
    "  2. The 22% / 14% ratios are the GUD-201 fallback, not the method. Table 12 (floor areas, beds, pupils) replaces them when data arrives.",
    "  3. The NCSI projection never levels off (691k by 2100). The 'ultimate' case must come from land capacity — see Saturation.",
]
for i, t in enumerate(lines, start=4):
    ws.cell(row=i, column=2, value=t).font = F_B if not t.startswith("  ") else F_B
ws["B4"].font = F_B
widths(ws, {"A": 2, "B": 130})

# ================================================================ Inputs
ws = wb.create_sheet("Inputs")
ws["B2"] = "Inputs — every number the calculation uses"
ws["B2"].font = F_TITLE
ws["B3"] = "Yellow cells are the ones to edit. Nothing else in the workbook is typed in twice."
ws["B3"].font = F_NOTE
header(ws, 5, ["Parameter", "Value", "Unit", "Source", "Status"], col0=2)
rows = [
    ("LPCD",       "Domestic consumption (Adh Dhahirah)", 164,   "l/c/d",           "GUD-201 Tab 11, G1-p60", "Standard"),
    ("ND_RATIO",   "Non-domestic ratio",                  0.22,  "of domestic",     "GUD-201 Tab 11, G1-p60", "FALLBACK — Table 12 land-use rates replace this when floor areas / counts arrive"),
    ("GOV_RATIO",  "Governmental ratio",                  0.14,  "of domestic",     "GUD-201 Tab 11, G1-p60", "FALLBACK — same as above"),
    ("RET_DOM",    "Return rate, domestic",               0.85,  "-",               "GUD-201 Tab 19, G1-p71", "Standard"),
    ("RET_NDG",    "Return rate, non-dom + gov",          0.54,  "-",               "GUD-201 Tab 19, G1-p71", "Standard"),
    ("INFIL_KM",   "Infiltration, new networks",          720,   "L/d per km",      "GUD-201 G1-p72",         "CONFLICT — R0 workbook uses 10% of flow; needs an NWS ruling"),
    ("PF_CAP",     "Peak factor cap",                     5.0,   "-",               "GUD-201 G1-p72",         "Standard"),
    ("STP_MARGIN", "STP design margin",                   0.10,  "on Qadf",         "GUD-201 G1-p73",         "Standard"),
    ("NET_KM",     "Sewer network length",                194,   "km",              "W2 trunk total",         "PROVISIONAL — laterals not included; effect is small (~140 m3/d)"),
    ("OR_RATE",    "Occupancy rate",                      6.0,   "persons/dwelling","[GAP-5]",                "UNCONFIRMED — colleague suggests 4.9; get officially from NCSI. See Saturation sheet"),
    ("DWELL_PLOT", "Dwellings per plot",                  1.0,   "dwellings/plot",  "[GAP-5]",                "UNCONFIRMED — pairs with occupancy: measured product is ~7 persons per built plot"),
    ("POP_PLOTS",  "Population-basis plots",              46633, "plots",           "W3 A1 (cadastre)",       "Grows if MoHUP subdivides or releases land"),
]
r = 6
for name, label, val, unit, src, status in rows:
    ws.cell(row=r, column=2, value=label).font = F_B
    c = ws.cell(row=r, column=3, value=val)
    c.font, c.fill, c.border = F_IN, FILL_IN, BOX
    c.number_format = "0%" if name in ("ND_RATIO", "GOV_RATIO", "STP_MARGIN") \
        else ("0.00" if name in ("RET_DOM", "RET_NDG") else "#,##0.0" if name in ("PF_CAP", "OR_RATE", "DWELL_PLOT") else NUM)
    ws.cell(row=r, column=4, value=unit).font = F_B
    ws.cell(row=r, column=5, value=src).font = F_B
    ws.cell(row=r, column=6, value=status).font = F_NOTE
    wb.defined_names[name] = DefinedName(name, attr_text=f"Inputs!$C${r}")
    r += 1
widths(ws, {"A": 2, "B": 34, "C": 10, "D": 15, "E": 24, "F": 72})
ws.freeze_panes = "A6"

# ================================================================ Population
ws = wb.create_sheet("Population")
ws["B2"] = "Population — the 25 project towns"
ws["B2"].font = F_TITLE
for i, t in enumerate([
    "Source: NCSI wilayat forecast, split to settlements by their frozen census share ('Ibri Sewer Demand R0 2026 08 03.xlsx').",
    "NCSI tracks 239 settlements in Ibri wilayat; the 25 inside the project boundary are these towns (63% of the wilayat).",
    "Every town grows at the same wilayat rate (x1.77-1.79 over 2025-2050) — there is no town-specific growth in the data.",
], start=3):
    ws.cell(row=i, column=2, value=t).font = F_NOTE
header(ws, 6, ["Town", "Code"] + [str(y) for y in YEARS_SHOW], col0=2)
r = 7
for _, t in towns.iterrows():
    ws.cell(row=r, column=2, value=t.NAME_EN).font = F_B
    ws.cell(row=r, column=3, value=str(t.TOWN)).font = F_B
    for j, y in enumerate(YEARS_SHOW):
        c = ws.cell(row=r, column=4 + j, value=int(t[f"Pop_{y}"]))
        c.font, c.number_format, c.border = F_B, NUM, BOX
    r += 1
TOT_ROW = r
ws.cell(row=r, column=2, value="TOTAL").font = F_H
for j in range(len(YEARS_SHOW)):
    col = get_column_letter(4 + j)
    c = ws.cell(row=r, column=4 + j, value=f"=SUM({col}7:{col}{r-1})")
    c.font, c.fill, c.number_format, c.border = F_H, FILL_TOT, NUM, BOX
widths(ws, {"A": 2, "B": 20, "C": 12} | {get_column_letter(4 + j): 10 for j in range(len(YEARS_SHOW))})
ws.freeze_panes = "B7"

# ================================================================ Load
ws = wb.create_sheet("Load")
ws["B2"] = "Load — from population to peak flow, one row per model year"
ws["B2"].font = F_TITLE
ws["B3"] = ("Chain per GUD-201 §7 (G1-p58..73): domestic water -> non-dom + gov -> return rates -> "
            "infiltration -> Qadf -> Peltier peak factor -> peak flow -> +10% STP margin.")
ws["B3"].font = F_NOTE
hdr = ["Year", "Population", "Domestic water m³/d", "Non-domestic m³/d", "Governmental m³/d",
       "Total water m³/d", "WW domestic m³/d", "WW non-dom+gov m³/d", "Infiltration m³/d",
       "Qadf m³/d", "Qm L/s", "Peak factor", "Peak flow m³/d", "STP design m³/d (+margin)"]
header(ws, 5, hdr, col0=2)
r = 6
for j, y in enumerate(YEARS_SHOW):
    popref = f"Population!{get_column_letter(4 + j)}${TOT_ROW}"
    ws.cell(row=r, column=2, value=str(y)).font = F_H
    cells = [
        (3, f"={popref}", F_LINK, NUM),
        (4, "=C{r}*LPCD/1000", F_B, NUM),
        (5, "=D{r}*ND_RATIO", F_B, NUM),
        (6, "=D{r}*GOV_RATIO", F_B, NUM),
        (7, "=D{r}+E{r}+F{r}", F_B, NUM),
        (8, "=D{r}*RET_DOM", F_B, NUM),
        (9, "=(E{r}+F{r})*RET_NDG", F_B, NUM),
        (10, "=NET_KM*INFIL_KM/1000", F_B, NUM),
        (11, "=H{r}+I{r}+J{r}", F_H, NUM),
        (12, "=K{r}/86.4", F_B, "#,##0.0"),
        (13, "=MIN(1.5+1/SQRT(L{r}),PF_CAP)", F_B, "0.00"),
        (14, "=K{r}*M{r}", F_B, NUM),
        (15, "=K{r}*(1+STP_MARGIN)", F_H, NUM),
    ]
    for col, f, fnt, fmt in cells:
        c = ws.cell(row=r, column=col, value=f.format(r=r))
        c.font, c.number_format, c.border = fnt, fmt, BOX
    r += 1
ws.cell(row=r + 1, column=2, value=("Note: Qadf here excludes tanker deliveries (~17% of STP inflow, G1-p73) — "
                                    "they go to the STP directly, not through the network.")).font = F_NOTE
widths(ws, {"A": 2, "B": 7} | {get_column_letter(c): 13 for c in range(3, 16)})
ws.freeze_panes = "C6"

# ================================================================ By Town 2055
ws = wb.create_sheet("By Town 2055")
ws["B2"] = "Load by town at 2055 — the pivotal model year"
ws["B2"].font = F_TITLE
for i, t in enumerate([
    "Domestic load follows each town's population. The non-domestic + governmental volume is NOT spread per person:",
    "it is placed on each town's non-residential plot area (commercial, governmental, mosque, industrial — from MoH_Plots).",
    "That is what GUD-201 §7.3.2-7.3.3 intend, and it moves load between towns without changing the total (see A7 note).",
], start=3):
    ws.cell(row=i, column=2, value=t).font = F_NOTE
header(ws, 6, ["Town", "Population 2055", "Domestic water m³/d", "Non-res plot area m²",
               "Share of non-res area", "ND+Gov allocated m³/d", "Qadf m³/d (excl. infil.)",
               "Peak factor", "Peak flow m³/d"], col0=2)
r = 7
N = len(towns)
for i, (_, t) in enumerate(towns.iterrows()):
    ws.cell(row=r, column=2, value=t.NAME_EN).font = F_B
    c = ws.cell(row=r, column=3, value=f"=Population!H{7+i}")     # H = 2055
    c.font, c.number_format, c.border = F_LINK, NUM, BOX
    for col, f, fmt in [
        (4, f"=C{r}*LPCD/1000", NUM),
        (5, int(t.nonres_m2), NUM),
        (6, f"=E{r}/SUM($E$7:$E${6+N})", "0.0%"),
        (7, f"=SUM($D$7:$D${6+N})*(ND_RATIO+GOV_RATIO)*F{r}", NUM),
        (8, f"=D{r}*RET_DOM+G{r}*RET_NDG", NUM),
        (9, f"=MIN(1.5+1/SQRT(H{r}/86.4),PF_CAP)", "0.00"),
        (10, f"=H{r}*I{r}", NUM),
    ]:
        c = ws.cell(row=r, column=col, value=f)
        c.font, c.number_format, c.border = F_B, fmt, BOX
    r += 1
ws.cell(row=r, column=2, value="TOTAL").font = F_H
for col in (3, 4, 5, 7, 8, 10):
    L = get_column_letter(col)
    c = ws.cell(row=r, column=col, value=f"=SUM({L}7:{L}{r-1})")
    c.font, c.fill, c.number_format, c.border = F_H, FILL_TOT, NUM, BOX
ws.cell(row=r + 2, column=2, value=("Non-res area source: MoH_Plots class v4 within town boundaries (W3 A7). "
                                    "AT TAYYIB's 218 ha is mostly the industrial estate — Table 12 data will "
                                    "re-price it properly. Peak factors apply at each town's own outlet.")).font = F_NOTE
widths(ws, {"A": 2, "B": 20} | {get_column_letter(c): 14 for c in range(3, 11)})
ws.freeze_panes = "C7"

# ================================================================ Saturation
ws = wb.create_sheet("Saturation")
ws["B2"] = "Saturation — how many people the plots can hold, and when we get there"
ws["B2"].font = F_TITLE
for i, t in enumerate([
    "Ceiling = population-basis plots x dwellings per plot x occupancy rate.",
    "The projection (right-hand table) keeps growing at ~2.3%/yr and never levels off — so the ceiling, not the",
    "projection, defines 'ultimate'. Measured today: ~7.0 persons per built plot (NCSI 2025 pop ÷ built plots).",
    "If occupancy is really 4.9, dwellings per plot must be ~1.2-1.4 for the sums to match the census.",
], start=3):
    ws.cell(row=i, column=2, value=t).font = F_NOTE
header(ws, 8, ["Occupancy rate", "Ceiling population", "Year the projection reaches it",
               "Ultimate Qadf m³/d", "STP design m³/d (+margin)"], col0=2)
YRS_COL, TOT_COL = "H", "I"
first, last = 9, 9 + len(ALL_YEARS) - 1
for k, orv in enumerate([4.9, 5.5, 6.0]):
    r = 9 + k
    c = ws.cell(row=r, column=2, value=orv)
    c.font, c.fill, c.border, c.number_format = F_IN, FILL_IN, BOX, "0.0"
    for col, f, fmt in [
        (3, f"=POP_PLOTS*DWELL_PLOT*B{r}", NUM),
        (4, f'=IFERROR(INDEX(${YRS_COL}${first}:${YRS_COL}${last},'
            f'MATCH(C{r},${TOT_COL}${first}:${TOT_COL}${last},1)+1),">2100")', "0"),
        (5, f"=C{r}*LPCD/1000*(RET_DOM+(ND_RATIO+GOV_RATIO)*RET_NDG)+NET_KM*INFIL_KM/1000", NUM),
        (6, f"=E{r}*(1+STP_MARGIN)", NUM),
    ]:
        c = ws.cell(row=r, column=col, value=f)
        c.font, c.number_format, c.border = F_B, fmt, BOX
ws.cell(row=13, column=2, value=("W2 used OR 6.0 -> ultimate Qadf ~49,700 m³/d. At 4.9 the land fills by ~2054 — "
                                 "inside the design horizon, which turns 2055 into a saturation case.")).font = F_NOTE
# projection totals, 2023-2100
header(ws, 8, ["Year", "Projection total"], col0=8)
yr_tot = {y: int(towns[f"Pop_{y}"].sum()) for y in ALL_YEARS}
for i, y in enumerate(ALL_YEARS):
    rr = first + i
    a = ws.cell(row=rr, column=8, value=y)
    b = ws.cell(row=rr, column=9, value=yr_tot[y])
    a.font = b.font = F_B
    b.number_format = NUM
ws.cell(row=first - 1, column=8).comment = None
widths(ws, {"A": 2, "B": 15, "C": 16, "D": 26, "E": 17, "F": 20, "G": 3, "H": 8, "I": 14})
ws.freeze_panes = "A9"

wb.save(OUT)
print(f"wrote {OUT}")
