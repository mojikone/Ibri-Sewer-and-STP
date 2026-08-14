# -*- coding: utf-8 -*-
"""Settlement land-capacity ceiling vs R0 projection.
Zones: Final_Boundary_IBRI.kmz 'Towns Boundary selection' polygons (WGS84 -> 32640).
Plots: MoH_Plots (61,272). Ceiling = pop-basis plots x OR 6.0 [GAP-5].
Projection: R0 workbook 'Project Pop Settlements' (2023-2100)."""
import zipfile, re, json, sys
import xml.etree.ElementTree as ET
import shapefile
from shapely.geometry import shape, Polygon, MultiPolygon, Point
from shapely.strtree import STRtree
from shapely.ops import transform as shp_transform
from rasterio.warp import transform as rio_transform
import openpyxl

KMZ = r"D:\Mojtaba\Renardet\2621 Ibri Sewer STP\Data\Received\2621\inception report - R0\Final_Boundary_IBRI.kmz"
PLOTS = r"D:\Mojtaba\Renardet\2621 Ibri Sewer STP\Data\MoHUP_DATA\MoH_Plots.shp"
XLSX = r"D:\Mojtaba\Renardet\2621 Ibri Sewer STP\Data\Received\2621\inception report - R0\Ibri Sewer Demand R0 2026 08 03.xlsx"
OUT = r"C:\Users\mojtaba\AppData\Local\Temp\claude\D--Mojtaba-Renardet-2621-Ibri-Sewer-STP-Hydraulic-Claude\413b098f-38f9-4a93-9820-2b7c34af7f5e\scratchpad"
OR_ = 6.0
NONRES = {"زراعى", "مسجد", "حكومي", "تجاري", "صناعي"}  # pure non-residential classes

# ---- 1. settlement polygons from kmz ----
ns = {"k": "http://www.opengis.net/kml/2.2"}
kml = zipfile.ZipFile(KMZ).read("doc.kml").decode("utf-8", errors="replace")
root = ET.fromstring(kml)
zones = {}
for pm in root.iter("{http://www.opengis.net/kml/2.2}Placemark"):
    nm_el = pm.find("k:name", ns)
    if nm_el is None: continue
    nm = nm_el.text.strip().upper()
    polys = []
    for poly in pm.iter("{http://www.opengis.net/kml/2.2}Polygon"):
        co = poly.find(".//k:outerBoundaryIs/k:LinearRing/k:coordinates", ns)
        if co is None: continue
        pts = []
        for tok in co.text.split():
            parts = tok.split(",")
            if len(parts) >= 2:
                pts.append((float(parts[0]), float(parts[1])))
        if len(pts) >= 4: polys.append(Polygon(pts))
    if polys:
        g = MultiPolygon(polys) if len(polys) > 1 else polys[0]
        zones[nm] = zones[nm].union(g) if nm in zones else g
print("zones with polygons:", len(zones), sorted(zones)[:30])

# transform to 32640
def to_utm(geom):
    def f(xs, ys, zs=None):
        x2, y2 = rio_transform("EPSG:4326", "EPSG:32640", list(xs), list(ys))
        return (x2, y2)
    return shp_transform(lambda x, y, z=None: f([x] if isinstance(x, float) else x, [y] if isinstance(y, float) else y) if False else None, geom)

# shapely transform needs vectorized func
def to_utm2(geom):
    def f(x, y, z=None):
        xs = list(x) if hasattr(x, "__iter__") else [x]
        ys = list(y) if hasattr(y, "__iter__") else [y]
        x2, y2 = rio_transform("EPSG:4326", "EPSG:32640", xs, ys)
        return (x2, y2)
    return shp_transform(f, geom)

zones = {nm: to_utm2(g) for nm, g in zones.items()}
areas = {nm: g.area / 1e6 for nm, g in zones.items()}

# ---- 2. plots -> centroids, classify ----
r = shapefile.Reader(PLOTS)
fields = [f[0] for f in r.fields[1:]]
i_lu = fields.index("LANDUSE")
cents = []
kinds = []
for sr in r.iterShapeRecords():
    try:
        g = shape(sr.shape.__geo_interface__)
        c = g.representative_point() if not g.is_empty else None
    except Exception:
        c = None
    if c is None: continue
    lu = str(sr.record[i_lu] or "").strip()
    kinds.append("nonres" if lu in NONRES else "popbasis")
    cents.append(c)
print("plot centroids:", len(cents))

tree = STRtree(cents)
res = {}
for nm, g in zones.items():
    idx = tree.query(g, predicate="contains")
    n_all = len(idx)
    n_pop = sum(1 for i in idx if kinds[i] == "popbasis")
    res[nm] = {"plots": n_all, "pop_basis": n_pop, "ceiling": n_pop * OR_, "area_km2": round(areas[nm], 1)}

# ---- 3. R0 projections ----
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Project Pop Settlements"]
hdr = [c.value for c in ws[1]]
yc = {int(str(h).replace("Pop ", "")): i for i, h in enumerate(hdr) if h and str(h).startswith("Pop ")}
proj = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[1]: continue
    nm = str(row[1]).strip().upper()
    proj[nm] = {y: float(row[i] or 0) for y, i in yc.items()}

# ---- 4. combine ----
rows = []
for nm, d in sorted(res.items(), key=lambda kv: -kv[1]["ceiling"]):
    p = proj.get(nm)
    if p:
        p24, p55, p2100 = p.get(2024, 0), p.get(2055, 0), p.get(2100, 0)
        cross = None
        for y in sorted(p):
            if p[y] >= d["ceiling"]:
                cross = y; break
        util55 = 100 * p55 / d["ceiling"] if d["ceiling"] else None
    else:
        p24 = p55 = p2100 = None; cross = None; util55 = None
    rows.append({"name": nm, **d, "pop2024": p24, "pop2055": p55, "pop2100": p2100,
                 "cross_year": cross, "util2055_pct": round(util55, 1) if util55 else None})
json.dump(rows, open(OUT + r"\zone_capacity.json", "w"))
print(f"{'settlement':<18}{'km2':>6}{'plots':>8}{'popbase':>8}{'ceil':>9}{'p2024':>9}{'p2055':>9}{'p2100':>9}{'cross':>7}{'u55%':>7}")
for d in rows:
    print(f"{d['name']:<18}{d['area_km2']:>6}{d['plots']:>8}{d['pop_basis']:>8}{d['ceiling']:>9.0f}"
          f"{(d['pop2024'] or 0):>9.0f}{(d['pop2055'] or 0):>9.0f}{(d['pop2100'] or 0):>9.0f}"
          f"{str(d['cross_year']):>7}{str(d['util2055_pct']):>7}")
tot_ceil = sum(d["ceiling"] for d in rows)
print("TOTAL ceiling (27 zones):", round(tot_ceil))
